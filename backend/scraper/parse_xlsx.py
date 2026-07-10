"""Download and parse guineematin results workbooks (BAC, BEPC, CEE all use the
same .xlsx shape when available — .xlsx is preferred over PDF, see discover.py).
"""

import io

import httpx
import openpyxl

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)


def download_xlsx(url: str) -> bytes:
    resp = httpx.get(url, headers={"User-Agent": USER_AGENT}, timeout=120.0, follow_redirects=True)
    resp.raise_for_status()
    return resp.content


def _clean(value) -> str:
    return value.strip() if isinstance(value, str) else (value or "")


def parse_results_xlsx(xlsx_bytes: bytes, examen: str):
    """
    Yield normalized row dicts from a results workbook.

    Each workbook has one sheet per track (e.g. BAC: one combined sheet with
    all profils incl. Franco-Arabe; BEPC/CEE: EG + FA sheets). The first
    column differs by exam: a real profil code for BAC (SS/SM/SE/SS-FA/SE-FA),
    a region (IRE for BEPC, DPE for CEE — not currently stored, see AGENTS.md)
    for BEPC/CEE, which have no filiere so every row gets the GENERAL sentinel.
    Some header/values have stray leading/trailing whitespace.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row is None or len(row) < 8:
                continue

            first_col, rang, ex, nom_complet, centre, pv, origine, mention = row[:8]

            if not pv or not str(pv).strip().isdigit():
                continue

            if examen == "BAC":
                profil = _clean(first_col)
                region = None
            else:
                # BEPC/CEE have no filiere — first_col is a region (IRE/DPE),
                # kept as part of the dedup key since PV is only unique within
                # a region for these two exams, not nationally (see ingest.py).
                profil = "GENERAL"
                region = _clean(first_col)

            yield {
                "profil": profil,
                "region": region,
                "rang": rang,
                "ex": ex,
                "nom_complet": _clean(nom_complet),
                "centre": _clean(centre),
                "pv": pv,
                "origine": _clean(origine),
                "mention": _clean(mention),
            }
